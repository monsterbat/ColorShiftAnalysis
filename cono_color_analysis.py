#!/usr/bin/env python
# -*- coding: utf-8 -*-\
# update : 2025/03/04

import os
import sys
import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime



###########################################
#            ColorShift_Analysis          #
###########################################

N = 5  # é€™è£¡å¯ä»¥æ”¹æˆ 10ï¼Œä»£è¡¨è¦ä¸€æ¬¡è™•ç†å¹¾å° Uã€V

# å»ºç«‹ 2 å€‹åˆ—è¡¨ï¼Œåˆ†åˆ¥å„²å­˜å¤šç­† Uã€V æª”æ¡ˆè·¯å¾‘
u_txt_paths = ["" for _ in range(N)]
v_txt_paths = ["" for _ in range(N)]
u_file_labels = []
v_file_labels = []

def choose_u_file(i):
    """é¸æ“‡ç¬¬ i å° U æª”æ¡ˆ"""
    global u_txt_paths
    file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
    if file_path:
        u_txt_paths[i] = file_path
        u_file_labels[i].config(text=os.path.basename(file_path))
    else:
        u_txt_paths[i] = ""
        u_file_labels[i].config(text="No file selected")

def choose_v_file(i):
    """é¸æ“‡ç¬¬ i å° V æª”æ¡ˆ"""
    global v_txt_paths
    file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
    if file_path:
        v_txt_paths[i] = file_path
        v_file_labels[i].config(text=os.path.basename(file_path))
    else:
        v_txt_paths[i] = ""
        v_file_labels[i].config(text="No file selected")



def load_txt_data(file_path):
    """
    å¾ TXT æ–‡ä»¶ä¸­è®€å–æ•¸æ“šï¼Œå¾ç¬¬ 4 è¡Œé–‹å§‹è§£æï¼Œä¸¦å¿½ç•¥éæ•¸å­—è¡Œ
    å›å‚³ NumPy é™£åˆ—
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()[3:]  # å¾ç¬¬ 4 è¡Œé–‹å§‹è®€å–
        data = []
        for line in lines:
            try:
                row = list(map(float, line.split()))
                if row:
                    data.append(row)
            except ValueError:
                continue  # å¿½ç•¥ç„¡æ³•è½‰æ›çš„è¡Œ
        return np.array(data)
    except Exception as e:
        print(f"è®€å–æª”æ¡ˆæ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        sys.exit(1)

def modify_data_formula(ws, rows, cols, original_ws):
    """
    åˆ©ç”¨ openpyxl ç›´æ¥è¨ˆç®— Modify å·¥ä½œè¡¨çš„æ•¸å€¼ï¼Œä¸ä½¿ç”¨ Excel å…¬å¼
    ä»¥ Original!A1 ç‚ºåŸºæº–ï¼Œå°‡æ¯å€‹å„²å­˜æ ¼çš„å€¼åšå·®
    """
    base_value = original_ws["A1"].value  # å–å¾— Original!A1 çš„æ•¸å€¼
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            col_letter = get_column_letter(c)
            original_value = original_ws[f"{col_letter}{r}"].value
            if original_value is not None:
                ws[f"{col_letter}{r}"] = original_value - base_value
            else:
                ws[f"{col_letter}{r}"] = None

def calculate_data_formula(ws, modify_ws):
    """
    åœ¨ Calculator å·¥ä½œè¡¨ä¸­å¡«å…¥æ•¸æ“šï¼š
      - A3~A143 ç”± -70 éå¢è‡³ 70
      - è¨­å®šæ¨™é¡Œï¼ˆå„æ¬„ä½ä¸åŒçš„åƒæ•¸ï¼‰
      - ä¾æ“š Modify å·¥ä½œè¡¨è³‡æ–™å¡«å…¥å…¶å®ƒæ¬„ä½ï¼ŒåŒ…å«éƒ¨åˆ†åè½‰æ’åˆ—èˆ‡è¤‡è£½
    """
    # å¡«å…¥ A3 ~ A143
    value = -70
    for r in range(3, 144):
        ws[f"A{r}"] = value
        value += 1

    # è¨­å®šæ¨™é¡Œ
    headers = {
        "B": [0, "0~70"], "C": [180, "0~70"], "D": [180, "-70~0"], "E": ["0~180", "FIN"],
        "F": [45, "0~70"], "G": [225, "0~70"], "H": [225, "-70~0"], "I": ["45~225", "FIN"],
        "J": [90, "0~70"], "K": [270, "0~70"], "L": [270, "-70~0"], "M": ["90~270", "FIN"],
        "N": [135, "0~70"], "O": [315, "0~70"], "P": [315, "-70~0"], "Q": ["135~315", "FIN"]
    }
    for col, (h1, h2) in headers.items():
        ws[f"{col}1"] = h1
        ws[f"{col}2"] = h2

    # å¡«å…¥ B3 ~ B73ï¼šæ ¹æ“š Modify å·¥ä½œè¡¨ä¸­æŒ‡å®šå„²å­˜æ ¼çš„æ•¸å€¼
    for i in range(71):
        ws[f"B{3+i}"] = modify_ws[f"A{1+i}"].value
        ws[f"C{3+i}"] = modify_ws[f"FY{1+i}"].value
        ws[f"F{3+i}"] = modify_ws[f"AT{1+i}"].value
        ws[f"G{3+i}"] = modify_ws[f"HR{1+i}"].value
        ws[f"J{3+i}"] = modify_ws[f"CM{1+i}"].value
        ws[f"K{3+i}"] = modify_ws[f"JK{1+i}"].value
        ws[f"N{3+i}"] = modify_ws[f"EF{1+i}"].value
        ws[f"O{3+i}"] = modify_ws[f"LD{1+i}"].value

    # å¡«å…¥ D3 ~ D73 (å°‡ C3~C73 åè½‰)
    for i in range(71):
        ws[f"D{3+i}"] = ws[f"C{73-i}"].value

    # å¡«å…¥ H3 ~ H73 (å°‡ G3~G73 åè½‰)
    for i in range(71):
        ws[f"H{3+i}"] = ws[f"G{73-i}"].value

    # å¡«å…¥ L3 ~ L73 (å°‡ K3~K73 åè½‰)
    for i in range(71):
        ws[f"L{3+i}"] = ws[f"K{73-i}"].value

    # å¡«å…¥ P3 ~ P73 (å°‡ O3~O73 åè½‰)
    for i in range(71):
        ws[f"P{3+i}"] = ws[f"O{73-i}"].value

    # å¡«å…¥ E3 ~ E72, I3 ~ I72, M3 ~ M72, Q3 ~ Q72 (è¤‡è£½éƒ¨åˆ†æ¬„ä½)
    for i in range(70):
        ws[f"E{3+i}"] = ws[f"D{3+i}"].value
        ws[f"I{3+i}"] = ws[f"H{3+i}"].value
        ws[f"M{3+i}"] = ws[f"L{3+i}"].value
        ws[f"Q{3+i}"] = ws[f"P{3+i}"].value

    # å¡«å…¥ E73 ~ E143 (å°‡ B3~B73 è¤‡è£½)
    for i in range(71):
        ws[f"E{73+i}"] = ws[f"B{3+i}"].value
        ws[f"I{73+i}"] = ws[f"F{3+i}"].value
        ws[f"M{73+i}"] = ws[f"J{3+i}"].value
        ws[f"Q{73+i}"] = ws[f"N{3+i}"].value

def result_data_formula(ws, calculate_ws):
    """
    åœ¨ Result å·¥ä½œè¡¨ä¸­å¡«å…¥æ•¸å€¼ï¼Œæ•´ç† Calculator å·¥ä½œè¡¨çš„æ•¸æ“šï¼š
      - A3~A143 ç‚ºæ–¹å‘ (-70 åˆ° 70)
      - å…¶é¤˜æ¬„ä½ä¾æ“š Calculator å·¥ä½œè¡¨è³‡æ–™å¡«å…¥
    """
    ws["A2"] = "direction"
    value = -70
    for r in range(3, 144):
        ws[f"A{r}"] = value
        value += 1

    ws["B2"] = "0~180"
    ws["C2"] = "45~225"
    ws["D2"] = "90~270"
    ws["E2"] = "135~315"

    for i in range(3, 144):
        ws[f"B{i}"] = calculate_ws[f"E{i}"].value
        ws[f"C{i}"] = calculate_ws[f"I{i}"].value
        ws[f"D{i}"] = calculate_ws[f"M{i}"].value
        ws[f"E{i}"] = calculate_ws[f"Q{i}"].value

def save_to_excel(original, txt_path):
    """
    å°‡è®€å–çš„åŸå§‹æ•¸æ“šå¯«å…¥ Excelï¼ˆå« Originalã€Modifyã€Calculateã€Result å››å€‹å·¥ä½œè¡¨ï¼‰
    Excel æª”åæ ¼å¼ï¼šResult_åŸæª”å_æ—¥æœŸæ™‚é–“.xlsx
    å›å‚³å„²å­˜å¾Œçš„ Excel æª”æ¡ˆè·¯å¾‘
    """
    original_filename = os.path.splitext(os.path.basename(txt_path))[0]
    timestamp = datetime.now().strftime("%Y%m%d%H%M")
    output_path = f"Result_{original_filename}_{timestamp}.xlsx"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Original"
    for r_idx, row in enumerate(original, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws1.cell(row=r_idx, column=c_idx, value=value)
    
    ws2 = wb.create_sheet("Modify")
    modify_data_formula(ws2, original.shape[0], original.shape[1], ws1)
    
    ws3 = wb.create_sheet("Calculate")
    calculate_data_formula(ws3, ws2)
    
    ws4 = wb.create_sheet("Result")
    result_data_formula(ws4, ws3)

    wb.save(output_path)
    print(f"Excel æª”æ¡ˆå·²ç”Ÿæˆï¼Œæª”åç‚º {output_path}")
    return output_path

###########################################
#                Combine                  #
###########################################

def read_excel_data(file_path):
    """
    è®€å– Excel æª”æ¡ˆä¸­çš„ 'Result' å·¥ä½œè¡¨ï¼Œå–å‡º B3 åˆ° E143 çš„æ•¸æ“š
    å›å‚³ä¸€å€‹ NumPy é™£åˆ—
    """
    if not file_path:
        return None
    try:
        xls = pd.ExcelFile(file_path)
        print(f"å¯ç”¨çš„å·¥ä½œè¡¨: {xls.sheet_names}")
        if 'Result' not in xls.sheet_names:
            print(f"éŒ¯èª¤: '{file_path}' ä¸åŒ…å« 'Result' å·¥ä½œè¡¨")
            return None
        df = pd.read_excel(file_path, sheet_name='Result', header=None)
        print("ğŸ“Š DataFrame é è¦½ï¼ˆå‰10è¡Œï¼‰ï¼š")
        print(df.head(10))
        print(f"DataFrame å¤§å°: {df.shape}")
        if df.shape[1] < 5:
            print("âš ï¸ è­¦å‘Š: Excel æª”æ¡ˆçš„åˆ—æ•¸å°‘æ–¼ 5ï¼Œå¯èƒ½æ•¸æ“šæ²’è®€å®Œæ•´ï¼")
        # å–å‡º B3 åˆ° E143 (æ³¨æ„ï¼špandas çš„ index å¾ 0 é–‹å§‹ï¼Œæ‰€ä»¥ row=2~142, col=1~4)
        data = df.iloc[2:143, 1:5].values  
        return data
    except Exception as e:
        print(f"è®€å– Excel æª”æ¡ˆæ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        return None

def calculate_ijkl_values(u_data, v_data):
    """
    åˆ©ç”¨ U èˆ‡ V æ•¸æ“šè¨ˆç®— Iã€Jã€Kã€L å€¼
    è¨ˆç®—å…¬å¼ï¼šsqrt( U^2 + V^2 )
    å›å‚³åŒ…å« I, J, K, L å››æ¬„çš„ DataFrame
    """
    if u_data is None or v_data is None:
        print("âŒ éŒ¯èª¤: æ²’æœ‰æœ‰æ•ˆçš„ U æˆ– V æ•¸æ“š")
        return None
    try:
        u_array = np.array(u_data, dtype=float)
        v_array = np.array(v_data, dtype=float)
    except Exception as e:
        print("âŒ è½‰æ›æ•¸æ“šæ™‚ç™¼ç”ŸéŒ¯èª¤:", e)
        return None

    if u_array.shape[1] < 4 or v_array.shape[1] < 4:
        print("âŒ éŒ¯èª¤: U æˆ– V æ•¸æ“šçš„åˆ—æ•¸ä¸è¶³")
        return None

    I_values = np.sqrt(u_array[:, 0] ** 2 + v_array[:, 0] ** 2)
    J_values = np.sqrt(u_array[:, 1] ** 2 + v_array[:, 1] ** 2)
    K_values = np.sqrt(u_array[:, 2] ** 2 + v_array[:, 2] ** 2)
    L_values = np.sqrt(u_array[:, 3] ** 2 + v_array[:, 3] ** 2)

    ijkl_df = pd.DataFrame({
        'I': I_values,
        'J': J_values,
        'K': K_values,
        'L': L_values
    })
    return ijkl_df

def create_new_excel(u_data, v_data, final_filename="output.xlsx"):
    """
    åˆ©ç”¨ U èˆ‡ V çš„æ•¸æ“šå‰µå»ºä¸€å€‹æ–° Excel æª”æ¡ˆ
    ...
    """
    if u_data is None or v_data is None:
        print("âŒ éŒ¯èª¤: æ²’æœ‰è®€å–åˆ°æœ‰æ•ˆæ•¸æ“š")
        return
    try:
        print("ğŸ” U è³‡æ–™å°ºå¯¸:", np.shape(u_data))
        print("ğŸ” V è³‡æ–™å°ºå¯¸:", np.shape(v_data))

        u_df = pd.DataFrame(u_data, columns=['A', 'B', 'C', 'D'])
        v_df = pd.DataFrame(v_data, columns=['E', 'F', 'G', 'H'])

        ijkl_df = calculate_ijkl_values(u_data, v_data)
        if ijkl_df is None:
            print("âŒ éŒ¯èª¤: ç„¡æ³•è¨ˆç®— Iã€Jã€Kã€L æ•¸æ“š")
            return

        with pd.ExcelWriter(final_filename, engine='openpyxl') as writer:
            u_df.to_excel(writer, index=False, header=False, startrow=3, startcol=0)
            v_df.to_excel(writer, index=False, header=False, startrow=3, startcol=4)
            ijkl_df.to_excel(writer, index=False, header=False, startrow=3, startcol=8)

        print(f"âœ… åˆä½µçµæœå·²å„²å­˜è‡³ {final_filename}")
    except Exception as e:
        print("âŒ ç”¢ç”Ÿåˆä½µ Excel æ™‚ç™¼ç”ŸéŒ¯èª¤:", e)




###########################################
#                ä¸»ç¨‹å¼                  #
###########################################

def main():
    root = tk.Tk()
    root.title("é¸å–å¤šç­† U èˆ‡ V æª”æ¡ˆ")

    # é€™è£¡è¦ä½¿ç”¨å…¨åŸŸçš„ u_file_labelsã€v_file_labels
    global u_file_labels, v_file_labels

    for i in range(N):
        # æ¯ä¸€ã€Œå°ã€ç”¨ä¸€å€‹ row_frame åŒ…èµ·ä¾†
        row_frame = tk.Frame(root)
        row_frame.pack(fill="x", padx=10, pady=5)

        # å·¦é‚Š (U)
        u_frame = tk.LabelFrame(row_frame, text=f"U (ç¬¬{i+1}çµ„)", padx=10, pady=10)
        u_frame.pack(side="left", fill="x", expand=True)

        # å»ºç«‹ Label, Button
        u_label = tk.Label(u_frame, text="No file selected")
        u_label.pack(pady=5)
        u_file_labels.append(u_label)  # å­˜åˆ°åˆ—è¡¨

        u_btn = tk.Button(u_frame, text="Select U File", command=lambda idx=i: choose_u_file(idx))
        u_btn.pack(pady=5)

        # å³é‚Š (V)
        v_frame = tk.LabelFrame(row_frame, text=f"V (ç¬¬{i+1}çµ„)", padx=10, pady=10)
        v_frame.pack(side="right", fill="x", expand=True)

        v_label = tk.Label(v_frame, text="No file selected")
        v_label.pack(pady=5)
        v_file_labels.append(v_label)

        v_btn = tk.Button(v_frame, text="Select V File", command=lambda idx=i: choose_v_file(idx))
        v_btn.pack(pady=5)

    # æœ€åº•éƒ¨ Combine æŒ‰éˆ•
    def start_processing():
        # æŒ‰ä¸‹å¾Œï¼Œé€ä¸€æª¢æŸ¥æ¯ä¸€çµ„ Uã€V
        for i in range(N):
            u_file = u_txt_paths[i]
            v_file = v_txt_paths[i]
            if u_file and v_file:
                process_files(u_file, v_file)
            else:
                print(f"ç¬¬{i+1}çµ„æœªé¸å–å®Œæ•´ï¼Œè·³é...")

        print("æ‰€æœ‰çµ„åˆ¥éƒ½è™•ç†å®Œæˆã€‚")

    combine_btn = tk.Button(root, text="Combine", command=start_processing)
    combine_btn.pack(side="bottom", pady=10)

    root.mainloop()


def process_files(u_file, v_file):
    analysis_excel_paths = []
    for txt_path in [u_file, v_file]:
        print("æ­£åœ¨è™•ç†æª”æ¡ˆï¼š", txt_path)
        original_data = load_txt_data(txt_path)
        excel_path = save_to_excel(original_data, txt_path)
        analysis_excel_paths.append(excel_path)

    print("é–‹å§‹åˆä½µåˆ†æçµæœ...")
    u_excel = analysis_excel_paths[0]
    v_excel = analysis_excel_paths[1]
    u_data = read_excel_data(u_excel)
    v_data = read_excel_data(v_excel)
    if u_data is None or v_data is None:
        print("ç„¡æ³•è®€å–åˆ†æçµæœé€²è¡Œåˆä½µã€‚")
        return

    # æ–°å¢ï¼šæ ¹æ“š U æª”åï¼Œå®šç¾©æœ€çµ‚æª”å
    u_basename = os.path.splitext(os.path.basename(u_file))[0]
    timestamp = datetime.now().strftime("%Y%m%d%H%M")
    final_filename = f"{u_basename}_result_{timestamp}.xlsx"

    create_new_excel(u_data, v_data, final_filename)
    print(f"ç¬¬ {u_basename} çµ„åˆä½µå®Œæˆï¼Œç”¢ç”Ÿæª”æ¡ˆï¼š{final_filename}")


if __name__ == "__main__":
    main()