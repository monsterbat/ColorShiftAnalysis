import pandas as pd
import os
import numpy as np
import sys
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def select_files():
    """彈出檔案選擇對話框，讓使用者選擇多個 TXT 檔案"""
    root = tk.Tk()
    root.withdraw()
    file_paths = filedialog.askopenfilenames(filetypes=[("Text files", "*.txt")])
    return file_paths  # 回傳多個檔案的路徑

def load_txt_data(file_path):
    """從 TXT 文件中讀取數據，從第 4 行開始解析，並忽略非數字行"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()[3:]  # 從第 4 行開始讀取
        
        data = []
        for line in lines:
            try:
                row = list(map(float, line.split()))  # 嘗試將該行轉換為 float
                if row:  # 確保行內有數據
                    data.append(row)
            except ValueError:
                continue  # 忽略無法轉換的行
        
        return np.array(data)
    except Exception as e:
        print(f"讀取檔案時發生錯誤: {e}")
        sys.exit(1)

def modify_data_formula(ws, rows, cols, original_ws):
    """直接在 Python 計算 Modify 分頁的數值，並寫入 Excel（不使用公式）"""
    base_value = original_ws["A1"].value  # 取得 Original!A1 的數值作為基準

    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            col_letter = get_column_letter(c)
            original_value = original_ws[f"{col_letter}{r}"].value  # 取得 Original 的數值
            if original_value is not None:
                ws[f"{col_letter}{r}"] = original_value - base_value  # **直接在 Python 計算數值**
            else:
                ws[f"{col_letter}{r}"] = None  # **如果原值是空的，就設為 None**
def calculate_data_formula(ws, modify_ws):
    """在 Excel 中填入 Calculator 分頁的數值"""
    # A3 ~ A143: -70 開始遞增到 70
    value = -70
    for r in range(3, 144):
        ws[f"A{r}"] = value
        value += 1

    # 設定標題
    headers = {
        "B": [0, "0~70"], "C": [180, "0~70"], "D": [180, "-70~0"], "E": ["0~180", "FIN"],
        "F": [45, "0~70"], "G": [225, "0~70"], "H": [225, "-70~0"], "I": ["45~225", "FIN"],
        "J": [90, "0~70"], "K": [270, "0~70"], "L": [270, "-70~0"], "M": ["90~270", "FIN"],
        "N": [135, "0~70"], "O": [315, "0~70"], "P": [315, "-70~0"], "Q": ["135~315", "FIN"]
    }
    
    for col, (h1, h2) in headers.items():
        ws[f"{col}1"] = h1
        ws[f"{col}2"] = h2

    # 填入 B3 ~ B73
    for i in range(71):
        ws[f"B{3+i}"] = modify_ws[f"A{1+i}"].value
        ws[f"C{3+i}"] = modify_ws[f"FY{1+i}"].value
        ws[f"F{3+i}"] = modify_ws[f"AT{1+i}"].value
        ws[f"G{3+i}"] = modify_ws[f"HR{1+i}"].value
        ws[f"J{3+i}"] = modify_ws[f"CM{1+i}"].value
        ws[f"K{3+i}"] = modify_ws[f"JK{1+i}"].value
        ws[f"N{3+i}"] = modify_ws[f"EF{1+i}"].value
        ws[f"O{3+i}"] = modify_ws[f"LD{1+i}"].value

    # 填入 D3 ~ D73 (C3 ~ C73 反轉)
    for i in range(71):
        ws[f"D{3+i}"] = ws[f"C{73-i}"].value

    # 填入 H3 ~ H73 (G3 ~ G73 反轉)
    for i in range(71):
        ws[f"H{3+i}"] = ws[f"G{73-i}"].value

    # 填入 L3 ~ L73 (K3 ~ K73 反轉)
    for i in range(71):
        ws[f"L{3+i}"] = ws[f"K{73-i}"].value

    # 填入 P3 ~ P73 (O3 ~ O73 反轉)
    for i in range(71):
        ws[f"P{3+i}"] = ws[f"O{73-i}"].value

    # 填入 E3 ~ E72, I3 ~ I72, M3 ~ M72, Q3 ~ Q72
    for i in range(70):
        ws[f"E{3+i}"] = ws[f"D{3+i}"].value
        ws[f"I{3+i}"] = ws[f"H{3+i}"].value
        ws[f"M{3+i}"] = ws[f"L{3+i}"].value
        ws[f"Q{3+i}"] = ws[f"P{3+i}"].value

    # 填入 E73 ~ E143 (B3 ~ B73)
    for i in range(71):
        ws[f"E{73+i}"] = ws[f"B{3+i}"].value
        ws[f"I{73+i}"] = ws[f"F{3+i}"].value
        ws[f"M{73+i}"] = ws[f"J{3+i}"].value
        ws[f"Q{73+i}"] = ws[f"N{3+i}"].value

def result_data_formula(ws, calculate_ws):
    """在 Excel 中填入 Result 分頁的數值，總整理 Calculate 分頁的數據"""
    ws["A2"] = "direction"
    
    # A3 ~ A143: -70 開始遞增到 70
    value = -70
    for r in range(3, 144):
        ws[f"A{r}"] = value
        value += 1

    # 設定標題
    ws["B2"] = "0~180"
    ws["C2"] = "45~225"
    ws["D2"] = "90~270"
    ws["E2"] = "135~315"

    # 填入 B3 ~ B143
    for i in range(3, 144):
        ws[f"B{i}"] = calculate_ws[f"E{i}"].value
        ws[f"C{i}"] = calculate_ws[f"I{i}"].value
        ws[f"D{i}"] = calculate_ws[f"M{i}"].value
        ws[f"E{i}"] = calculate_ws[f"Q{i}"].value

def save_to_excel(original, txt_path):
    # 取得原始檔案名稱（不含副檔名）
    original_filename = os.path.splitext(os.path.basename(txt_path))[0]
    
    # 取得當前時間
    timestamp = datetime.now().strftime("%Y%m%d%H%M")

    # 設定 Excel 檔名格式為 Result_原檔名_日期時間.xlsx
    output_path = f"Result_{original_filename}_{timestamp}.xlsx"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Original"
    for r_idx, row in enumerate(original, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws1.cell(row=r_idx, column=c_idx, value=value)
    
    ws2 = wb.create_sheet("Modify")
    modify_data_formula(ws2, original.shape[0], original.shape[1], ws1)
    
    ws3 = wb.create_sheet("Calculate")
    calculate_data_formula(ws3, ws2)
    
    ws4 = wb.create_sheet("Result")
    result_data_formula(ws4,ws3)

    wb.save(output_path)
    print(f"Excel 檔案已生成，檔名為 {output_path}")

def main():
    txt_paths = select_files()  # 改為多選
    if not txt_paths:
        print("未選擇檔案，程式結束。")
        return
    
    for txt_path in txt_paths:
        print("正在處理檔案：", txt_path)
        original_data = load_txt_data(txt_path)
        save_to_excel(original_data, txt_path)  # 針對每個檔案執行

if __name__ == "__main__":
    main()
